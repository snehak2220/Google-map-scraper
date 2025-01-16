from selenium import webdriver
import time
import csv
import requests
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import chromedriver_autoinstaller
from selenium.common.exceptions import TimeoutException
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# Headers for HTTP requests
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36'
}

# Excel sheet name
sheet_name = 'jwelleries_palakkad_list.xlsx'

# Function to create Excel sheet with headlines
def xl_sheet_headlines(sheet_name=sheet_name):
    wb = Workbook()
    ws = wb.active
    headlines = ['url', 'name', 'address', 'website', 'phone', 'category', 'email']
    ws.append(headlines)
    wb.save(sheet_name)

xl_sheet_headlines()

# Function to write data to an existing Excel sheet
def xl_write(data_write, sheet_name=sheet_name):
    wb = load_workbook(sheet_name)
    work_sheet = wb.active  # Get active sheet
    work_sheet.append(data_write)
    wb.save(sheet_name)

# Function to set up the Chrome driver
def driver_define():
    print('Chromedriver Installing')
    driver_path = chromedriver_autoinstaller.install()

    print('Chrome Browser Opening')
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    s = Service(driver_path)
    driver = webdriver.Chrome(service=s, options=options)
    return driver

# Function to get email from a URL
def get_email(url):
    domain = url.split('//')[-1].replace('www.', '').split('/')[0]
    url_gen = f'http://www.skymem.info/srch?q={domain}'
    response = requests.get(url_gen, headers=headers)
    soup = BeautifulSoup(response.text, 'lxml')
    email_list = re.findall(r"href=\"\/srch\?q=(.*?@.*)\">", str(soup))
    email = [line for line in email_list if domain in line][0]

    return email

# Main driver and scraping logic
driver = driver_define()
urls_filename = 'jwellery_urls.txt'
urls = [line.strip('\n') for line in open(urls_filename).readlines()]

for url in urls:
    driver.get(url)

    print('--------------------------')

    try:
        name = WebDriverWait(driver, 25).until(
            EC.visibility_of_element_located((By.XPATH, '//div[@class="specific-class"]/h1'))
        ).text
    except TimeoutException as e:
        print(f"TimeoutException: {e}")
        name = 'N/A'
    except Exception as e:
        print(f"Error occurred: {e}")
        name = 'N/A'

    time.sleep(5)

    try:
        address = driver.find_element(By.XPATH, '//button[@data-item-id="address"]').text
    except:
        address = 'N/A'

    try:
        website = driver.find_element(By.CSS_SELECTOR, 'a[aria-label^="Website:"]').get_attribute('href')
    except:
        website = 'N/A'

    try:
        phone = driver.find_element(By.CSS_SELECTOR, 'button[aria-label*="Phone:"]').text
    except:
        phone = 'N/A'

    try:
        category = driver.find_element(By.CSS_SELECTOR, '[jsaction="pane.rating.category"]').text
    except:
        category = 'N/A'

    email = 'N/A'
    try:
        if len(website) > 3:
            email = get_email(website)
    except:
        email = 'N/A'

    print(f"name : {name}")
    print(f"address : {address}")
    print(f"website:", website)
    print(f"phone:", phone)
    print("category:", category)
    print("email:", email)

    write_data = [url, name, address, website, phone, category, email]

    xl_write(write_data)

# Close the driver after processing all URLs
driver.quit()
